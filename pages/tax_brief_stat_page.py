from dataclasses import dataclass
from nicegui import ui,events, app, run
from components import inputs, tables,cards
from typing import Optional, Any
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from utils import global_vars as g
from dao.tax_approval_dao import TaxApprovalDao

@dataclass
class SearchCondition:
    select_year: str = ""
search_condition = SearchCondition()

async def show_tax_brief_stat_page():
    with ui.row().classes('w-full h-[80px] px-[20px] mt-0 place-content-start') \
        .style('background-color: #FFFFFF !important; border-radius: 10px;'):
        with ui.row().classes('h-full items-center'):
            curyear = datetime.now().year
            years_value = [f'{year}年' for year in range(curyear, 2020, -1)]
            inputs.selection_w40(years_value, years_value[0], False, None) \
                .bind_value_to(search_condition, 'select_year')
        with ui.row().classes('h-full items-center'):
            ui.button('刷新', icon='img:/static/images/refresh@2x.png', on_click=on_search) \
                .classes('w-25 rounded-md text-white') \
                .style('background-color: #6C96FB !important')
            ui.button('导出', icon='file_download', on_click=export_tax_brief) \
                .classes('w-25 rounded-md text-white') \
                .style('background-color: #4CAF50 !important')
    out_grid_columns = 13*6 + 2
    with ui.row().classes('w-full px-[20px] mt-2 place-content-start flex flex-col') \
        .style('height: calc(100dvh - 80px); background-color: #FFFFFF !important; border-radius: 10px;'):
        with ui.scroll_area().classes('w-full h-full flex-1') \
            .style('background-color: #FFFFFF !important; border-radius: 10px; overflow-x: auto;'):
                app.storage.client['tax_brief_grid'] = ui.grid(columns=out_grid_columns).classes('w-full h-full gap-0 items-center')\
                    .style('grid-template-columns: repeat(80,60px)')
        # 初始化分页状态（按公司分页）
        app.storage.client.setdefault('tax_brief_paging', {'page': 1, 'page_size': 20, 'total': 0})
        with ui.row().classes('w-full items-center justify-center gap-2 mt-2'):
            async def go_first(*_) -> None:
                await on_search(1)

            async def go_prev(*_) -> None:
                await on_search(max(1, app.storage.client['tax_brief_paging']['page'] - 1))

            async def go_next(*_) -> None:
                await on_search(app.storage.client['tax_brief_paging']['page'] + 1)

            async def go_last(*_) -> None:
                p = app.storage.client['tax_brief_paging']
                last_page = max(1, (p.get('total', 0) + p.get('page_size', 20) - 1) // p.get('page_size', 20))
                await on_search(last_page)

            first_btn = ui.button('首页', on_click=go_first)
            prev_btn = ui.button('上一页', on_click=go_prev)
            app.storage.client['tax_brief_page_label'] = ui.label('')
            next_btn = ui.button('下一页', on_click=go_next)
            last_btn = ui.button('尾页', on_click=go_last)
            app.storage.client['tax_brief_first_btn'] = first_btn
            app.storage.client['tax_brief_prev_btn'] = prev_btn
            app.storage.client['tax_brief_next_btn'] = next_btn
            app.storage.client['tax_brief_last_btn'] = last_btn

        await on_search()
                        
            

async def on_search(page: int = 1) -> None:
    if 'tax_brief_grid' not in app.storage.client:
        return
    app.storage.client['tax_brief_grid'].clear()
    app.storage.client['tax_brief_rows'] = []
    paging = app.storage.client.setdefault('tax_brief_paging', {'page': 1, 'page_size': 20, 'total': 0})
    try:
        page = int(page)
    except Exception:
        page = 1
    if page < 1:
        page = 1
    paging['page'] = page

    result, company_info = g.query_company_name_company()
    if result is False:
        ui.notify('查询公司信息失败')
        return
    # 按内部公司筛选并分页
    company_list = [c for c in company_info.values() if c.brief_name and len(c.brief_name) > 0 and c.type == 1]
    total = len(company_list)
    page_size = paging.get('page_size', 20)
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    total_pages = max(1, total_pages)
    if page > total_pages:
        page = total_pages
        paging['page'] = page
    paging['total'] = total
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_companies = company_list[start_idx:end_idx]
    with app.storage.client['tax_brief_grid']:
        ui.label('').classes('col-span-2 text-lx font-bold')
        ui.label('年度合计').classes('col-span-6 flex justify-center text-lx font-bold border p-1 border-gray-300')
        for i in range(1, 13):
            ui.label(f'{i}月').classes('col-span-6 flex justify-center text-lx font-bold border p-1 border-gray-300')
        ui.label('').classes('col-span-2 text-lx font-bold')
        for i in range(1, 14):
            ui.label('增值税').classes('text-sm flex justify-center font-normal border-l border-b p-1 border-gray-300')
            ui.label('附加').classes('text-sm flex justify-center font-normal border-b p-1 border-gray-300')
            ui.label('印花税').classes('text-sm flex justify-center font-normal border-b p-1 border-gray-300')
            ui.label('所得税').classes('text-sm flex justify-center font-normal border-b p-1 border-gray-300')
            ui.label('其他税').classes('text-sm flex justify-center font-normal border-b p-1 border-gray-300')
            ui.label('合计').classes('text-sm flex justify-center font-normal border-r border-b p-1 border-gray-300')
        def do_search() -> tuple[bool, list[dict], str]:
            rows = []
            for company in page_companies:
                year_dict = {'year_tax_value': 0.0, 'year_attach_value': 0.0, 'year_stamp_value': 0.0, 'year_income_value': 0.0, 'year_other_value': 0.0, 'year_total_value': 0.0}
                month_dict = {}
                for i in range(1, 13):
                    month_dict[f'{i}'] = {'month_tax_value': 0.0, 'month_attach_value': 0.0, 'month_stamp_value': 0.0, 'month_income_value': 0.0, 'month_other_value': 0.0, 'month_total_value': 0.0}
                    result, tax_values = g.my_db.query_tax_approval_by_period_date(company.id, f'{search_condition.select_year[:-1]}-{i:02d}')
                    if result and tax_values and len(tax_values) > 0:
                        for tax_value in tax_values:
                            dao = TaxApprovalDao()
                            dao.from_db(tax_value)
                            match dao.tax_type:
                                case '增值税':
                                    month_dict[f'{i}']['month_tax_value'] += dao.paid_in_money
                                case '地方教育附加':
                                    month_dict[f'{i}']['month_attach_value'] += dao.paid_in_money
                                case '教育费附加':
                                    month_dict[f'{i}']['month_attach_value'] += dao.paid_in_money
                                case '城市维护建设税':
                                    month_dict[f'{i}']['month_attach_value'] += dao.paid_in_money
                                case '印花税':
                                    month_dict[f'{i}']['month_stamp_value'] += dao.paid_in_money
                                case '企业所得税':
                                    month_dict[f'{i}']['month_income_value'] += dao.paid_in_money
                                case _:
                                    month_dict[f'{i}']['month_other_value'] += dao.paid_in_money
                            month_dict[f'{i}']['month_total_value'] += dao.paid_in_money
                        year_dict['year_tax_value'] += month_dict[f'{i}']['month_tax_value']
                        year_dict['year_attach_value'] += month_dict[f'{i}']['month_attach_value']
                        year_dict['year_stamp_value'] += month_dict[f'{i}']['month_stamp_value']
                        year_dict['year_income_value'] += month_dict[f'{i}']['month_income_value']
                        year_dict['year_other_value'] += month_dict[f'{i}']['month_other_value']
                        year_dict['year_total_value'] += month_dict[f'{i}']['month_total_value']
                values = {'company': company, 'year_dict': year_dict}
                for i in range(1, 13):
                    values[f'month_dict_{i}'] = month_dict[f'{i}']
                rows.append({company.brief_name: values})
            return True, rows, ''
            
        refresh_dialog = g.show_refresh_process("统计中，请稍候")
        success, rows, message = await run.io_bound(do_search)
        if not success:
            ui.notify(message or '查询记录失败')
        refresh_dialog.close()
        for row in rows:
            for company_name, values in row.items():
                ui.label(company_name).classes('col-span-2 text-sm flex justify-center font-normal border-r border-b p-1 border-gray-300')
                year_dict = values['year_dict']
                year_tax_value = year_dict['year_tax_value']
                year_attach_value = year_dict['year_attach_value']
                year_stamp_value = year_dict['year_stamp_value']
                year_income_value = year_dict['year_income_value']
                year_other_value = year_dict['year_other_value']
                year_total_value = year_dict['year_total_value']
                year_tax_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-l border-b p-1 border-gray-300')
                year_attach_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                year_stamp_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                year_income_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                year_other_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                year_total_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-r border-b p-1 border-gray-300')
                year_tax_value_label.set_text(g.format_currency(year_tax_value))
                year_attach_value_label.set_text(g.format_currency(year_attach_value))
                year_stamp_value_label.set_text(g.format_currency(year_stamp_value))
                year_income_value_label.set_text(g.format_currency(year_income_value))
                year_other_value_label.set_text(g.format_currency(year_other_value))
                year_total_value_label.set_text(g.format_currency(year_total_value))
                row_values = [company_name, year_tax_value, year_attach_value, year_stamp_value, year_income_value, year_other_value, year_total_value]
                for i in range(1, 13):
                    month_dict = values[f'month_dict_{i}']
                    month_tax_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-l border-b p-1 border-gray-300')
                    month_attach_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                    month_stamp_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                    month_income_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                    month_other_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-b p-1 border-gray-300')
                    month_total_value_label = ui.label('').classes('text-xs flex justify-center font-normal border-r border-b p-1 border-gray-300')
            
                    month_tax_value_label.set_text(g.format_currency(month_dict['month_tax_value']))
                    month_attach_value_label.set_text(g.format_currency(month_dict['month_attach_value']))
                    month_stamp_value_label.set_text(g.format_currency(month_dict['month_stamp_value']))
                    month_income_value_label.set_text(g.format_currency(month_dict['month_income_value']))
                    month_other_value_label.set_text(g.format_currency(month_dict['month_other_value']))
                    month_total_value_label.set_text(g.format_currency(month_dict['month_total_value']))

                    row_values.extend([month_dict['month_tax_value'], month_dict['month_attach_value'], month_dict['month_stamp_value'], month_dict['month_income_value'], month_dict['month_other_value'], month_dict['month_total_value']])
                app.storage.client['tax_brief_rows'].append(row_values)
    # 更新分页显示与按钮状态
    paging['total'] = total
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    total_pages = max(1, total_pages)
    app.storage.client['tax_brief_page_label'].set_text(f"第 {page} / {total_pages} 页，共 {total} 条")
    app.storage.client['tax_brief_first_btn'].disabled = (page <= 1)
    app.storage.client['tax_brief_prev_btn'].disabled = (page <= 1)
    app.storage.client['tax_brief_next_btn'].disabled = (page >= total_pages)
    app.storage.client['tax_brief_last_btn'].disabled = (page >= total_pages)
async def export_tax_brief():
    # 重新统计全部公司的数据并导出（不再只导出当前分页）
    result, company_info = g.query_company_name_company()
    if result is False:
        ui.notify('查询公司信息失败')
        return
    company_list = [c for c in company_info.values() if c.brief_name and len(c.brief_name) > 0 and c.type == 1]
    if not company_list:
        ui.notify('没有数据可导出')
        return

    def build_all_rows() -> tuple[bool, list, str]:
        rows = []
        for company in company_list:
            year_dict = {'year_tax_value': 0.0, 'year_attach_value': 0.0, 'year_stamp_value': 0.0, 'year_income_value': 0.0, 'year_other_value': 0.0, 'year_total_value': 0.0}
            month_dict = {}
            for i in range(1, 13):
                month_dict[f'{i}'] = {'month_tax_value': 0.0, 'month_attach_value': 0.0, 'month_stamp_value': 0.0, 'month_income_value': 0.0, 'month_other_value': 0.0, 'month_total_value': 0.0}
                res, tax_values = g.my_db.query_tax_approval_by_period_date(company.id, f'{search_condition.select_year[:-1]}-{i:02d}')
                if res and tax_values and len(tax_values) > 0:
                    for tax_value in tax_values:
                        dao = TaxApprovalDao()
                        dao.from_db(tax_value)
                        match dao.tax_type:
                            case '增值税':
                                month_dict[f'{i}']['month_tax_value'] += dao.paid_in_money
                            case '地方教育附加' | '教育费附加' | '城市维护建设税':
                                month_dict[f'{i}']['month_attach_value'] += dao.paid_in_money
                            case '印花税':
                                month_dict[f'{i}']['month_stamp_value'] += dao.paid_in_money
                            case '企业所得税':
                                month_dict[f'{i}']['month_income_value'] += dao.paid_in_money
                            case _:
                                month_dict[f'{i}']['month_other_value'] += dao.paid_in_money
                        month_dict[f'{i}']['month_total_value'] += dao.paid_in_money
                    year_dict['year_tax_value'] += month_dict[f'{i}']['month_tax_value']
                    year_dict['year_attach_value'] += month_dict[f'{i}']['month_attach_value']
                    year_dict['year_stamp_value'] += month_dict[f'{i}']['month_stamp_value']
                    year_dict['year_income_value'] += month_dict[f'{i}']['month_income_value']
                    year_dict['year_other_value'] += month_dict[f'{i}']['month_other_value']
                    year_dict['year_total_value'] += month_dict[f'{i}']['month_total_value']
            row_values = [company.brief_name, year_dict['year_tax_value'], year_dict['year_attach_value'], year_dict['year_stamp_value'], year_dict['year_income_value'], year_dict['year_other_value'], year_dict['year_total_value']]
            for i in range(1, 13):
                md = month_dict[f'{i}']
                row_values.extend([md['month_tax_value'], md['month_attach_value'], md['month_stamp_value'], md['month_income_value'], md['month_other_value'], md['month_total_value']])
            rows.append(row_values)
        return True, rows, ''

    refresh_dialog = g.show_refresh_process('导出中，请稍候')
    success, rows_all, message = await run.io_bound(build_all_rows)
    refresh_dialog.close()
    if not success or not rows_all:
        ui.notify(message or '没有数据可导出')
        return

    # 更新缓存（方便页面继续使用）
    app.storage.client['tax_brief_rows'] = rows_all

    columns = ['公司名称', '增值税', '附加', '印花税', '所得税', '其他税', '合计']
    for _ in range(1, 13):
        columns.extend(['增值税', '附加', '印花税', '所得税', '其他税', '合计'])
    df = pd.DataFrame(rows_all, columns=columns)
    fname = f'./static/完税汇总_{search_condition.select_year}.xlsx'
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)
        ws = writer.sheets['Sheet1']

        for i, col in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            values = df.iloc[:, i].astype(str).replace(['nan', 'None'], '')
            max_len = max(values.map(len).max(), len(str(col))) + 5
            ws.column_dimensions[column_letter].width = max_len

        ws.merge_cells('B1:G1')
        ws['B1'] = '年度合计'
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        for i in range(1, 13):
            start_col = 8 + (i - 1) * 6
            end_col = start_col + 5
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.value = f'{i}月'
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ui.download.file(fname)